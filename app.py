# app.py – Smart Diff Manager (CLEAN UI + Fixed Multi-Key Detection)

import streamlit as st
import pandas as pd
from difflib import SequenceMatcher
from io import BytesIO
import msoffcrypto
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from typing import List, Dict, Tuple, Optional
import re

DEFAULT_PASSWORD = "mypassword"

st.set_page_config(
    page_title="Smart Diff Manager",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ----------------------------------------------------------------------
# STYLES
# ----------------------------------------------------------------------
st.markdown(
    """
<style>
.dataframe-container { font-size: 14px; }
.small-text { font-size: 13px; opacity: 0.9; margin-left: 1rem; }
</style>
""",
    unsafe_allow_html=True,
)

# ----------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------

def decrypt_file(uploaded_file, password: Optional[str] = DEFAULT_PASSWORD) -> BytesIO:
    fb = BytesIO(uploaded_file.getvalue())
    try:
        pd.ExcelFile(fb)
        fb.seek(0)
        return fb
    except Exception:
        fb.seek(0)
        office = msoffcrypto.OfficeFile(fb)
        if not office.is_encrypted():
            fb.seek(0)
            return fb
        dec = BytesIO()
        try:
            office.load_key(password=password)
            office.decrypt(dec)
            dec.seek(0)
            return dec
        except Exception:
            fb.seek(0)
            return fb

# ----------------------------------------------------------------------
# NORMALIZATION UTILITIES
# ----------------------------------------------------------------------

def normalize_colname(name: str) -> str:
    """Normalize column names for comparison - remove special chars, lowercase"""
    return re.sub(r'[^a-z0-9]', '', str(name).lower().strip())


def normalize_logical(v):
    """
    Normalize various representations of boolean and numeric values for comparison.
    - TRUE/FALSE handling stays the same
    - Numeric values like 123.0 vs 123 are unified
    """
    if pd.isna(v) or v == "":
        return ""

    s = str(v).strip().upper()

    # Boolean normalization
    t = {"TRUE","T","YES","Y","1","1.0","CHECK","CHECKED","CHECKMARK","✓","✔","ON","ENABLED","ACTIVE"}
    f = {"FALSE","F","NO","N","0","0.0","CROSS","UNCHECKED","✗","✘","X","OFF","DISABLED","INACTIVE"}
    if s in t: return "TRUE"
    if s in f: return "FALSE"

    # --- Numeric normalization ---
    try:
        # Replace commas with dots just in case locale uses comma decimals
        sn = s.replace(",", ".")
        num = float(sn)
        # If integer-valued, cast to int string
        if abs(num - int(num)) < 1e-9:
            return str(int(num))
        # Otherwise, keep up to one decimal (remove trailing zeroes)
        return re.sub(r'\\.0+$', '', str(num))
    except Exception:
        pass  # not numeric, continue below

    return s


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize dataframe - fill NaN, convert to string, and normalize boolean values"""
    if df is None or df.empty:
        return pd.DataFrame()
    d = df.copy().fillna("")
    # Make sure column names are strings to avoid KeyError when numeric columns exist
    d.columns = d.columns.map(str)
    for c in d.columns:
        d[c] = d[c].astype(str).apply(normalize_logical)
    return d

# ----------------------------------------------------------------------
# SMART HEADER DETECTION - Find key columns ANYWHERE in the row
# ----------------------------------------------------------------------

def find_header_row_by_column_names(df_raw: pd.DataFrame, reference_columns: List[str]) -> int:
    if not reference_columns:
        return detect_header_row_heuristic(df_raw)
    normalized_ref = {normalize_colname(str(c)) for c in reference_columns if str(c).strip()}
    best_row = 0
    best_match_count = 0
    for i in range(min(15, len(df_raw))):
        row = df_raw.iloc[i]
        row_values = {
            normalize_colname(str(val))
            for val in row
            if pd.notna(val) and str(val).strip()
        }
        match_count = len(normalized_ref & row_values)
        if match_count >= max(2, len(normalized_ref) * 0.5):
            if match_count > best_match_count:
                best_match_count = match_count
                best_row = i
                if match_count >= len(normalized_ref) * 0.8:
                    return best_row
    if best_match_count >= 2:
        return best_row
    return detect_header_row_heuristic(df_raw)


def find_header_row_with_keys(df_raw: pd.DataFrame, key_columns: List[str]) -> int:
    if not key_columns:
        return detect_header_row_heuristic(df_raw)
    normalized_keys = {normalize_colname(k) for k in key_columns if k.strip()}
    for i in range(min(15, len(df_raw))):
        row = df_raw.iloc[i]
        row_values = []
        for val in row:
            if pd.notna(val) and str(val).strip():
                row_values.append(normalize_colname(str(val)))
        if normalized_keys & set(row_values):
            return i
    return detect_header_row_heuristic(df_raw)


def detect_header_row_heuristic(df_raw: pd.DataFrame) -> int:
    max_nonempty = df_raw.head(15).apply(lambda r: r.notna().sum(), axis=1).max()
    for i, row in df_raw.head(15).iterrows():
        filled = row.notna().sum()
        if filled < 2:
            continue
        non_null_values = [str(val) for val in row if pd.notna(val) and str(val).strip()]
        if len(non_null_values) < 2:
            continue
        filled_ratio = filled / max_nonempty if max_nonempty else 0
        text_ratio = sum(isinstance(x, str) and str(x).strip() for x in row) / max(1, filled)
        valid_indices = [idx for idx, val in enumerate(row) if pd.notna(val) and str(val).strip()]
        if valid_indices:
            span = max(valid_indices) - min(valid_indices) + 1
            contiguous = len(valid_indices) / span
        else:
            contiguous = 0
        if filled_ratio >= 0.4 and text_ratio >= 0.5 and contiguous > 0.5:
            return i
    return 0


def read_visible_sheets_with_header_detection(
    file_bytes: BytesIO,
    key_columns: List[str] = None,
    reference_headers: Dict[str, List[str]] = None
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, int]]:
    try:
        wb = load_workbook(file_bytes, data_only=True)
        visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        sheets = {}
        header_rows = {}
        for sh in visible_sheets:
            ws = wb[sh]
            visible_col_indices = []
            for col_idx in range(1, ws.max_column + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                col_dim = ws.column_dimensions.get(col_letter)
                is_visible = (col_dim is None) or (not col_dim.hidden)
                if is_visible:
                    visible_col_indices.append(col_idx - 1)
            if not visible_col_indices:
                visible_col_indices = list(range(ws.max_column))
            file_bytes.seek(0)
            df_raw = pd.read_excel(file_bytes, sheet_name=sh, header=None, engine="openpyxl")
            if visible_col_indices and len(visible_col_indices) < len(df_raw.columns):
                df_raw = df_raw.iloc[:, visible_col_indices]
            header_row = 0
            if reference_headers and sh in reference_headers:
                header_row = find_header_row_by_column_names(df_raw, reference_headers[sh])
            elif key_columns:
                header_row = find_header_row_with_keys(df_raw, key_columns)
            else:
                header_row = detect_header_row_heuristic(df_raw)
            file_bytes.seek(0)
            df_full = pd.read_excel(file_bytes, sheet_name=sh, header=header_row, engine="openpyxl")
            if visible_col_indices and len(visible_col_indices) < len(df_full.columns):
                df_main = df_full.iloc[:, visible_col_indices]
            else:
                df_main = df_full
            if isinstance(df_main.columns, pd.MultiIndex):
                df_main.columns = [
                    " ".join([str(c) for c in col if str(c) != "nan"]).strip()
                    for col in df_main.columns
                ]
            df_main = df_main.astype(str)
            sheets[sh] = df_main
            header_rows[sh] = header_row
        return sheets, header_rows
    except Exception as e:
        try:
            file_bytes.seek(0)
            all_sheets_raw = pd.read_excel(file_bytes, sheet_name=None, engine="openpyxl", header=None)
            sheets = {}
            header_rows = {}
            for sh, df_raw in all_sheets_raw.items():
                if reference_headers and sh in reference_headers:
                    header_row = find_header_row_by_column_names(df_raw, reference_headers[sh])
                elif key_columns:
                    header_row = find_header_row_with_keys(df_raw, key_columns)
                else:
                    header_row = detect_header_row_heuristic(df_raw)
                file_bytes.seek(0)
                df = pd.read_excel(file_bytes, sheet_name=sh, header=header_row, engine="openpyxl")
                df = df.astype(str)
                sheets[sh] = df
                header_rows[sh] = header_row
            return sheets, header_rows
        except Exception:
            file_bytes.seek(0)
            all_sheets = pd.read_excel(file_bytes, sheet_name=None, engine="openpyxl")
            for sh in all_sheets:
                all_sheets[sh] = all_sheets[sh].astype(str)
            return all_sheets, {}

# ----------------------------------------------------------------------
# KEY SELECTION - Find best key that exists in both files
# ----------------------------------------------------------------------

def find_best_valid_key(df1: pd.DataFrame, df2: pd.DataFrame, keys: List[str]) -> Tuple[List[str], str]:
    """
    Find the best valid key column that exists in both dataframes.
    Now supports partial matches like 'emplid_0' ↔ 'emplid'.
    """
    if not keys:
        return [], "No keys provided"

    def norm(s): 
        s = str(s).strip().lower()
        s = re.sub(r'(_\\d+)+$', '', s)  # remove trailing _0, _1 etc.
        return re.sub(r'[^a-z0-9]', '', s)

    norm1 = {norm(c): c for c in df1.columns}
    norm2 = {norm(c): c for c in df2.columns}

    best_key = None
    best_uniqueness = 0
    best_key_name = None

    for k in keys:
        nk = norm(k)
        # First try exact match
        if nk in norm1 and nk in norm2:
            actual_col = norm1[nk]
        else:
            # Try partial match (e.g. emplid matches emplid_0)
            candidates1 = [c for n, c in norm1.items() if nk in n or n in nk]
            candidates2 = [c for n, c in norm2.items() if nk in n or n in nk]
            if not (candidates1 and candidates2):
                continue
            actual_col = candidates1[0]  # pick first matching column

        try:
            unique1 = df1[actual_col].fillna("").astype(str).nunique()
            unique2 = df2[actual_col].fillna("").astype(str).nunique()
            uniqueness = (unique1 / max(1, len(df1)) + unique2 / max(1, len(df2))) / 2
            if uniqueness > best_uniqueness:
                best_uniqueness = uniqueness
                best_key = [actual_col]
                best_key_name = k
        except Exception:
            continue

    if best_key:
        return best_key, f"Key: '{best_key_name}' ({best_uniqueness:.0%} unique)"
    return [], "No valid keys found"

# ----------------------------------------------------------------------
# COMPARISON FUNCTIONS
# ----------------------------------------------------------------------

def compare_key_based(df1, df2, keys):
    valid_keys, key_desc = find_best_valid_key(df1, df2, keys)
    if not valid_keys:
        raise ValueError("No valid key columns found")
    common = list(df1.columns.intersection(df2.columns))
    df1f, df2f = df1[common].copy(), df2[common].copy()
    df1f = df1f.fillna("")
    df2f = df2f.fillna("")
    n1, n2 = normalize_df(df1f), normalize_df(df2f)
    n1["__key__"] = n1[valid_keys].astype(str).agg("||".join, axis=1)
    n2["__key__"] = n2[valid_keys].astype(str).agg("||".join, axis=1)
    df1f["__key__"], df2f["__key__"] = n1["__key__"], n2["__key__"]
    k1, k2 = set(n1["__key__"]), set(n2["__key__"])
    common_keys, added_keys, removed_keys = k1 & k2, k2 - k1, k1 - k2
    changed_old, changed_new = [], []
    for key in common_keys:
        r1 = n1.loc[n1["__key__"] == key]
        r2 = n2.loc[n2["__key__"] == key]
        if r1.empty or r2.empty:
            continue
        r1, r2 = r1.iloc[0], r2.iloc[0]
        if any(str(r1[c]) != str(r2[c]) for c in common if c not in valid_keys and c != "__key__"):
            changed_old.append(df1f.loc[df1f["__key__"] == key].iloc[0])
            changed_new.append(df2f.loc[df2f["__key__"] == key].iloc[0])
    co = pd.DataFrame(changed_old)[common] if changed_old else pd.DataFrame(columns=common)
    cn = pd.DataFrame(changed_new)[common] if changed_new else pd.DataFrame(columns=common)
    added = df2f.loc[df2f["__key__"].isin(added_keys), common].reset_index(drop=True) if added_keys else pd.DataFrame(columns=common)
    removed = df1f.loc[df1f["__key__"].isin(removed_keys), common].reset_index(drop=True) if removed_keys else pd.DataFrame(columns=common)
    return co.reset_index(drop=True), cn.reset_index(drop=True), added, removed, key_desc


def compare_keyless(df1, df2):
    n1, n2 = normalize_df(df1), normalize_df(df2)
    h1, h2 = pd.util.hash_pandas_object(n1, index=False), pd.util.hash_pandas_object(n2, index=False)
    added = df2.loc[~h2.isin(h1)].copy().reset_index(drop=True)
    removed = df1.loc[~h1.isin(h2)].copy().reset_index(drop=True)
    return added, removed


def detect_data_truncation(df1: pd.DataFrame, df2: pd.DataFrame) -> Tuple[bool, int, int]:
    old_non_empty = df1.apply(lambda row: row.notna().any(), axis=1).sum()
    new_non_empty = df2.apply(lambda row: row.notna().any(), axis=1).sum()
    threshold = max(5, old_non_empty * 0.1)
    is_truncated = (old_non_empty - new_non_empty) >= threshold
    return is_truncated, old_non_empty, new_non_empty

# ----------------------------------------------------------------------
# PREVIEW & EXPORT WITH HIGHLIGHTING
# ----------------------------------------------------------------------

def build_side_by_side_preview(changed_old: pd.DataFrame, changed_new: pd.DataFrame, keys: List[str]):
    common_cols = changed_old.columns.intersection(changed_new.columns).tolist()
    old_df = changed_old[common_cols].copy()
    new_df = changed_new[common_cols].copy()
    old_df_norm = normalize_df(old_df)
    new_df_norm = normalize_df(new_df)
    combined_data = []
    for idx in range(len(old_df)):
        row_data = {}
        for col in common_cols:
            old_val = str(old_df.iloc[idx][col])
            new_val = str(new_df.iloc[idx][col])
            row_data[f"{col} (Old)"] = old_val
            row_data[f"{col} (New)"] = new_val
        combined_data.append(row_data)
    combined_df = pd.DataFrame(combined_data)
    def highlight_changes(row):
        styles = []
        row_idx = row.name
        for i, col in enumerate(combined_df.columns):
            if col.endswith(" (Old)"):
                base_col = col[:-6]
                if base_col in old_df_norm.columns:
                    old_val_norm = str(old_df_norm.iloc[row_idx][base_col])
                    new_val_norm = str(new_df_norm.iloc[row_idx][base_col])
                    if old_val_norm != new_val_norm:
                        styles.append('background-color: #ffcccc; color: #000000')
                    else:
                        styles.append('background-color: #e8f5e9; color: #000000')
                else:
                    styles.append('')
            elif col.endswith(" (New)"):
                base_col = col[:-6]
                if base_col in new_df_norm.columns:
                    old_val_norm = str(old_df_norm.iloc[row_idx][base_col])
                    new_val_norm = str(new_df_norm.iloc[row_idx][base_col])
                    if old_val_norm != new_val_norm:
                        styles.append('background-color: #ccffcc; color: #000000')
                    else:
                        styles.append('background-color: #e8f5e9; color: #000000')
                else:
                    styles.append('')
            else:
                styles.append('')
        return styles
    styled_df = combined_df.style.apply(highlight_changes, axis=1)
    return styled_df


def style_added_rows(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    df = df.copy()
    # Ensure unique column names for Styler
    if not df.columns.is_unique:
        df.columns = [f"{col}_{i}" for i, col in enumerate(df.columns)]
    def highlight(row):
        return ['background-color: #e3f2fd; color: #000000' for _ in row]
    return df.style.apply(highlight, axis=1)


def style_removed_rows(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    df = df.copy()
    # Ensure unique column names for Styler
    if not df.columns.is_unique:
        df.columns = [f"{col}_{i}" for i, col in enumerate(df.columns)]
    def highlight(row):
        return ['background-color: #ffebee; color: #000000' for _ in row]
    return df.style.apply(highlight, axis=1)


def export_to_excel(changed_old: pd.DataFrame, changed_new: pd.DataFrame, keys: List[str]) -> BytesIO:
    wb = Workbook()
    ws_old = wb.active
    ws_old.title = "Old Values"
    ws_new = wb.create_sheet(title="New Values")
    old_highlight = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    new_highlight = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    changed_old_norm = normalize_df(changed_old)
    changed_new_norm = normalize_df(changed_new)
    for c, col_name in enumerate(changed_old.columns, 1):
        ws_old.cell(row=1, column=c, value=col_name)
    for r_idx in range(len(changed_old)):
        for c_idx, col_name in enumerate(changed_old.columns, 1):
            cell = ws_old.cell(row=r_idx + 2, column=c_idx, value=changed_old.iloc[r_idx][col_name])
            if (col_name not in keys and 
                str(changed_old_norm.iloc[r_idx][col_name]) != str(changed_new_norm.iloc[r_idx][col_name])):
                cell.fill = old_highlight
    for c, col_name in enumerate(changed_new.columns, 1):
        ws_new.cell(row=1, column=c, value=col_name)
    for r_idx in range(len(changed_new)):
        for c_idx, col_name in enumerate(changed_new.columns, 1):
            cell = ws_new.cell(row=r_idx + 2, column=c_idx, value=changed_new.iloc[r_idx][col_name])
            if (col_name not in keys and 
                str(changed_old_norm.iloc[r_idx][col_name]) != str(changed_new_norm.iloc[r_idx][col_name])):
                cell.fill = new_highlight
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ----------------------------------------------------------------------
# UI AND EXECUTION
# ----------------------------------------------------------------------
st.title("📊 Smart Diff Manager")
st.markdown("Compare Excel files with smart header detection and key-based matching.")

left = st.file_uploader("Upload **OLD** files", ["xlsx","xls"], True)
right = st.file_uploader("Upload **NEW** files", ["xlsx","xls"], True)
keys_str = st.text_input("Key Columns (comma-separated, e.g. ID, Name)", "")

if left and right:
    ignore_suffix = st.checkbox("Ignore suffix after last underscore", True)
    threshold = st.slider("Auto-match threshold", 0.5, 1.0, 0.85, 0.05)
    def comparable(n, ignore):
        n = n.rsplit(".", 1)[0]
        return n.rsplit("_", 1)[0].lower() if ignore and "_" in n else n.lower()
    matched = []
    usedL, usedR = set(), set()
    for lf in left:
        lcmp = comparable(lf.name, ignore_suffix)
        best, ratio = None, 0
        for rf in right:
            if rf.name in usedR:
                continue
            rcmp = comparable(rf.name, ignore_suffix)
            r = SequenceMatcher(None, lcmp, rcmp).ratio()
            if r >= threshold and r > ratio:
                best, ratio = rf, r
        if best:
            matched.append((lf, best))
            usedL.add(lf.name)
            usedR.add(best.name)
    st.write(f"**{len(matched)}** file pair(s) matched")
    if matched:
        with st.expander("📋 View matched pairs"):
            for lf, rf in matched:
                st.write(f"• `{lf.name}` ↔️ `{rf.name}`")
    if st.button("▶️ Run Comparison"):
        prog = st.progress(0)
        keys = [k.strip() for k in keys_str.split(",") if k.strip()]
        files_with_changes = []
        for i, (lf, rf) in enumerate(matched):
            try:
                decL, decR = decrypt_file(lf), decrypt_file(rf)
                shL, header_rows_old = read_visible_sheets_with_header_detection(decL, key_columns=keys)
                # ----------------------------------------------------------------------
                # Align NEW file headers to OLD file header rows and column names
                # ----------------------------------------------------------------------
                def _normalize_for_match(s: str) -> str:
                    s = str(s).strip().lower()
                    s = re.sub(r'(_\d+)+$', '', s)  # remove trailing _0/_1 suffixes
                    return re.sub(r'[^a-z0-9]', '', s)
                def align_new_columns_to_reference(new_cols: List[str], ref_cols: List[str]) -> List[str]:
                    new_cols = [str(c) for c in new_cols]
                    ref_norm_map = {i: _normalize_for_match(c) for i, c in enumerate(ref_cols)}
                    new_norm_map = {i: _normalize_for_match(c) for i, c in enumerate(new_cols)}
                    assigned_new_idx = set()
                    renamed = list(new_cols)
                    for ref_i, ref_norm in ref_norm_map.items():
                        for new_i, new_norm in new_norm_map.items():
                            if new_i in assigned_new_idx:
                                continue
                            if ref_norm and new_norm == ref_norm:
                                renamed[new_i] = ref_cols[ref_i]
                                assigned_new_idx.add(new_i)
                                break
                    for ref_i, ref_norm in ref_norm_map.items():
                        if not ref_norm:
                            continue
                        for new_i, new_norm in new_norm_map.items():
                            if new_i in assigned_new_idx:
                                continue
                            if ref_norm in new_norm or new_norm in ref_norm:
                                renamed[new_i] = ref_cols[ref_i]
                                assigned_new_idx.add(new_i)
                                break
                    final = []
                    seen = {}
                    for name in renamed:
                        base = str(name)
                        if base in seen:
                            seen[base] += 1
                            final_name = f"{base}_{seen[base]}"
                        else:
                            seen[base] = 0
                            final_name = base
                        final.append(final_name)
                    return final
                shR = {}
                header_rows_new = {}
                for sh in shL.keys():
                    hr_old = header_rows_old.get(sh, 0)
                    decR.seek(0)
                    df_new_raw = pd.read_excel(decR, sheet_name=sh, header=None, engine="openpyxl")
                    if hr_old < len(df_new_raw):
                        new_header = df_new_raw.iloc[hr_old].tolist()
                        df_new = df_new_raw.iloc[hr_old + 1:].copy()
                        df_new.columns = new_header
                    else:
                        df_new = pd.read_excel(decR, sheet_name=sh, header=0, engine="openpyxl")
                    try:
                        ref_cols = list(shL[sh].columns)
                        aligned_cols = align_new_columns_to_reference(list(df_new.columns), ref_cols)
                        df_new.columns = aligned_cols
                    except Exception:
                        df_new.columns = [str(c) for c in df_new.columns]
                    shR[sh] = df_new.astype(str)
                    header_rows_new[sh] = hr_old
            except Exception as e:
                st.error(f"❌ File pair {i+1}/{len(matched)}: Failed to read files: {e}")
                import traceback
                st.code(traceback.format_exc())
                prog.progress((i + 1) / len(matched))
                continue
            changed_sheets = []
            for sh in sorted(set(shL) & set(shR)):
                d1, d2 = shL[sh], shR[sh]
                use_key_based = False
                key_desc = ""
                if keys:
                    try:
                        valid_keys, key_desc = find_best_valid_key(d1, d2, keys)
                        use_key_based = len(valid_keys) > 0
                    except Exception:
                        use_key_based = False
                try:
                    if use_key_based:
                        co, cn, add, rem, key_desc = compare_key_based(d1, d2, keys)
                    else:
                        co = cn = pd.DataFrame()
                        add, rem = compare_keyless(d1, d2)
                except Exception:
                    co = cn = pd.DataFrame()
                    add, rem = compare_keyless(d1, d2)
                if not (cn.empty and add.empty and rem.empty):
                    changed_sheets.append((sh, co, cn, add, rem, use_key_based, key_desc))
            if changed_sheets:
                files_with_changes.append((i+1, lf.name, rf.name, changed_sheets))
            prog.progress((i + 1) / len(matched))
        if not files_with_changes:
            st.success(f"✅ All {len(matched)} file pair(s) are identical!")
        else:
            st.info(f"📊 Found differences in {len(files_with_changes)} out of {len(matched)} file pair(s)")
            for file_num, old_name, new_name, changed_sheets in files_with_changes:
                st.markdown(f"## 📂 File Pair {file_num}/{len(matched)}")
                st.markdown(f"**OLD:** `{old_name}`  \n**NEW:** `{new_name}`")
                for sheet_data in changed_sheets:
                    if len(sheet_data) == 10:
                        sh, co, cn, add, rem, is_key_based, key_desc, is_truncated, old_rows, new_rows = sheet_data
                    else:
                        sh, co, cn, add, rem, is_key_based, key_desc = sheet_data
                        is_truncated, old_rows, new_rows = False, 0, 0
                    summary_parts = []
                    if len(cn) > 0:
                        summary_parts.append(f"{len(cn)} changed")
                    if len(add) > 0:
                        summary_parts.append(f"{len(add)} added")
                    if len(rem) > 0:
                        summary_parts.append(f"{len(rem)} removed")
                    summary = ", ".join(summary_parts) if summary_parts else "differences detected"
                    with st.expander(f"📄 `{sh}` - {summary}", expanded=True):
                        if is_truncated:
                            st.error(f"⚠️ **Data Truncation Detected!**")
                            st.write(f"OLD file has **{old_rows} rows** but NEW file only has **{new_rows} rows**")
                            st.write(f"**{old_rows - new_rows} rows** of data are missing in the NEW file!")
                        if is_key_based and key_desc and "Key:" in key_desc:
                            st.caption(f"🔑 {key_desc}")
                        elif key_desc and "Row-based" in key_desc:
                            st.caption(f"ℹ️ {key_desc} (no matching key column found)")
                        if not cn.empty and is_key_based:
                            st.write("**🔄 Changed Rows**")
                            preview = build_side_by_side_preview(co, cn, keys)
                            st.dataframe(preview, width="stretch")
                            excel_file = export_to_excel(co, cn, keys)
                            st.download_button(
                                "📥 Download Excel",
                                excel_file,
                                file_name=f"Pair{file_num}_{sh}_changes.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{file_num}_{sh}"
                            )
                        elif not cn.empty:
                            st.write("**🔄 Changed Rows**")
                            st.dataframe(cn, width="stretch")
                        if not add.empty:
                            st.write("**➕ Added Rows**")
                            st.dataframe(style_added_rows(add), width="stretch")
                        if not rem.empty:
                            st.write("**➖ Removed Rows**")
                            st.dataframe(style_removed_rows(rem), width="stretch")
else:
    st.info("Please upload files to start comparison.")
