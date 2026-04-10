# app.py – Smart Diff Manager (Improved: UI/UX + Performance + Header Detection)

import streamlit as st
import pandas as pd
from difflib import SequenceMatcher
from io import BytesIO
import msoffcrypto
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from typing import List, Dict, Tuple, Optional
import re
import hashlib

DEFAULT_PASSWORD = "mypassword"

st.set_page_config(
    page_title="Smart Diff Manager",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* ── Page background ── */
.stApp {
    background: #0f1117;
    color: #e2e8f0;
}

/* ── Top banner ── */
.banner {
    background: linear-gradient(135deg, #1a1f2e 0%, #0f1117 60%, #1a2744 100%);
    border: 1px solid #2d3748;
    border-radius: 12px;
    padding: 2rem 2.5rem 1.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.banner::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(59,130,246,0.12) 0%, transparent 70%);
    pointer-events: none;
}
.banner h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.8rem;
    font-weight: 600;
    color: #f1f5f9;
    margin: 0 0 0.4rem;
    letter-spacing: -0.5px;
}
.banner p {
    color: #94a3b8;
    font-size: 0.9rem;
    margin: 0;
    font-weight: 300;
}
.banner .accent {
    color: #60a5fa;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    display: block;
    margin-bottom: 0.5rem;
}

/* ── Step cards ── */
.step-card {
    background: #1a1f2e;
    border: 1px solid #2d3748;
    border-radius: 10px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}
.step-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #60a5fa;
    margin-bottom: 0.6rem;
    font-weight: 600;
}

/* ── Metric pills ── */
.metric-row {
    display: flex;
    gap: 1rem;
    flex-wrap: wrap;
    margin: 1rem 0;
}
.metric-pill {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 8px;
    padding: 0.6rem 1.1rem;
    font-size: 0.85rem;
    color: #cbd5e1;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
.metric-pill .val {
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    font-size: 1.1rem;
}
.metric-pill.changed .val { color: #fbbf24; }
.metric-pill.added   .val { color: #34d399; }
.metric-pill.removed .val { color: #f87171; }
.metric-pill.info    .val { color: #60a5fa; }

/* ── Sheet header ── */
.sheet-header {
    display: flex;
    align-items: center;
    gap: 0.75rem;
    padding: 0.9rem 1.2rem;
    background: #1a1f2e;
    border-left: 3px solid #60a5fa;
    border-radius: 0 8px 8px 0;
    margin: 1.5rem 0 0.75rem;
}
.sheet-header .sheet-name {
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    font-size: 0.95rem;
    color: #e2e8f0;
}
.sheet-header .sheet-summary {
    color: #64748b;
    font-size: 0.82rem;
}

/* ── Diff section labels ── */
.diff-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    font-weight: 600;
    padding: 0.3rem 0.7rem;
    border-radius: 4px;
    display: inline-block;
    margin-bottom: 0.5rem;
}
.diff-label.changed { background: #451a03; color: #fbbf24; }
.diff-label.added   { background: #052e16; color: #34d399; }
.diff-label.removed { background: #450a0a; color: #f87171; }
.diff-label.key     { background: #172554; color: #93c5fd; }

/* ── File pair card ── */
.pair-card {
    background: #1a1f2e;
    border: 1px solid #2d3748;
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin-bottom: 0.5rem;
    display: flex;
    align-items: center;
    gap: 1rem;
    font-size: 0.85rem;
}
.pair-card .fname { font-family: 'IBM Plex Mono', monospace; color: #94a3b8; }
.pair-card .arrow { color: #3b82f6; font-size: 1rem; }

/* ── Warn / error banners ── */
.warn-box {
    background: #1c1007;
    border: 1px solid #92400e;
    border-radius: 8px;
    padding: 0.8rem 1.2rem;
    color: #fcd34d;
    font-size: 0.85rem;
    margin-bottom: 0.75rem;
}
.success-box {
    background: #052e16;
    border: 1px solid #166534;
    border-radius: 8px;
    padding: 1rem 1.5rem;
    color: #86efac;
    font-size: 0.95rem;
    text-align: center;
    font-family: 'IBM Plex Mono', monospace;
}

/* ── Streamlit overrides ── */
div[data-testid="stFileUploader"] {
    background: #1a1f2e;
    border: 1px dashed #334155;
    border-radius: 8px;
    padding: 0.5rem;
}
div[data-testid="stTextInput"] > div > div > input {
    background: #1e293b;
    border: 1px solid #334155;
    color: #e2e8f0;
    border-radius: 6px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
}
div[data-testid="stSlider"] { padding: 0.2rem 0; }
.stButton > button {
    background: #2563eb;
    color: white;
    border: none;
    border-radius: 8px;
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    letter-spacing: 0.5px;
    padding: 0.6rem 1.8rem;
    font-size: 0.9rem;
    transition: background 0.2s;
}
.stButton > button:hover { background: #1d4ed8; }
.stCheckbox label { color: #94a3b8; font-size: 0.85rem; }
.stExpander { border: 1px solid #2d3748 !important; border-radius: 8px !important; background: #1a1f2e; }
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────
# BANNER
# ─────────────────────────────────────────────
st.markdown(
    """
<div class="banner">
  <span class="accent">v2.0 · Excel Diff Tool</span>
  <h1>📊 Smart Diff Manager</h1>
  <p>Key-based comparison across Excel files with smart header detection and column alignment.</p>
</div>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────
# HELPER: stable file hash for caching
# ─────────────────────────────────────────────

def _file_hash(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


# ─────────────────────────────────────────────
# DECRYPTION (cached by content hash)
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def decrypt_file_cached(data: bytes, password: str = DEFAULT_PASSWORD) -> bytes:
    fb = BytesIO(data)
    try:
        pd.ExcelFile(fb)
        return data
    except Exception:
        pass
    fb.seek(0)
    try:
        office = msoffcrypto.OfficeFile(fb)
        if not office.is_encrypted():
            return data
        dec = BytesIO()
        office.load_key(password=password)
        office.decrypt(dec)
        return dec.getvalue()
    except Exception:
        return data


def decrypt_file(uploaded_file, password: str = DEFAULT_PASSWORD) -> BytesIO:
    decrypted = decrypt_file_cached(uploaded_file.getvalue(), password)
    return BytesIO(decrypted)


# ─────────────────────────────────────────────
# EAGER AUTO-DETECTION (cached, runs on upload)
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def detect_header_rows_for_file(file_bytes_raw: bytes) -> Dict[str, int]:
    """
    Run smart header detection on every sheet in a file immediately after upload.
    Returns {sheet_name: detected_row (1-indexed, for display)}.
    Cached by file content so it only runs once per unique file.
    """
    result: Dict[str, int] = {}
    try:
        wb = load_workbook(BytesIO(file_bytes_raw), data_only=True)
        all_raw = pd.read_excel(BytesIO(file_bytes_raw), sheet_name=None,
                                header=None, engine="openpyxl")
        for ws in wb.worksheets:
            sh = ws.title
            df_raw = all_raw.get(sh)
            if df_raw is None or df_raw.empty:
                result[sh] = 1
                continue
            detected = detect_header_row_heuristic(df_raw, ws=ws)
            result[sh] = detected + 1   # convert to 1-indexed for display
    except Exception:
        pass
    return result


def get_auto_detected_rows(uploaded_files: list) -> Dict[str, int]:
    """
    Merge auto-detected header rows across all uploaded files.
    If two files disagree on the same sheet, take the higher row number
    (more conservative — skips more potential title rows).
    """
    merged: Dict[str, int] = {}
    for uf in (uploaded_files or []):
        try:
            rows = detect_header_rows_for_file(uf.getvalue())
            uf.seek(0)
            for sh, row in rows.items():
                # Take the max (most conservative) if files disagree
                merged[sh] = max(merged.get(sh, 1), row)
        except Exception:
            try:
                uf.seek(0)
            except Exception:
                pass
    return merged


# ─────────────────────────────────────────────
# NORMALIZATION UTILITIES
# ─────────────────────────────────────────────

def normalize_colname(name: str) -> str:
    return re.sub(r'[^a-z0-9]', '', str(name).lower().strip())


# Boolean lookup sets
_BOOL_TRUE  = {"TRUE","T","YES","Y","1","1.0","CHECK","CHECKED","CHECKMARK","✓","✔","ON","ENABLED","ACTIVE"}
_BOOL_FALSE = {"FALSE","F","NO","N","0","0.0","CROSS","UNCHECKED","✗","✘","X","OFF","DISABLED","INACTIVE"}

def _normalize_cell(v: str) -> str:
    """Normalize a single string cell value."""
    if v == "" or v == "nan":
        return ""
    # Collapse all whitespace variants (newlines, tabs, non-breaking spaces,
    # multiple spaces) into a single space, then strip ends.
    # This prevents 'Current\nComp-Ratio' and 'Current Comp-Ratio' from
    # hashing differently when they look identical in a cell.
    s = re.sub(r'[\s\u00a0]+', ' ', v).strip().upper()
    if s in _BOOL_TRUE:
        return "TRUE"
    if s in _BOOL_FALSE:
        return "FALSE"
    try:
        num = float(s.replace(",", "."))
        return str(int(num)) if abs(num - int(num)) < 1e-9 else re.sub(r'\.0+$', '', str(num))
    except Exception:
        return s


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Vectorized normalization: fill NaN → stringify → normalize booleans/numerics.
    Much faster than per-cell apply() for large frames.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    d = df.copy()
    d.columns = d.columns.map(str)
    # Vectorized fillna + astype in one pass
    d = d.fillna("").astype(str)
    # Apply scalar function column-by-column (still faster than cell-by-cell)
    for col in d.columns:
        d[col] = d[col].map(_normalize_cell)
    return d


# ─────────────────────────────────────────────
# SMART HEADER DETECTION  (fixed index handling)
# ─────────────────────────────────────────────

def find_header_row_by_column_names(df_raw: pd.DataFrame, reference_columns: List[str]) -> int:
    if not reference_columns:
        return detect_header_row_heuristic(df_raw)
    normalized_ref = {normalize_colname(str(c)) for c in reference_columns if str(c).strip()}
    best_row = 0
    best_match_count = 0
    n_rows = min(15, len(df_raw))
    for i in range(n_rows):                          # ← use range, not iterrows index
        row = df_raw.iloc[i]
        row_values = {
            normalize_colname(str(val))
            for val in row
            if pd.notna(val) and str(val).strip()
        }
        match_count = len(normalized_ref & row_values)
        if match_count >= max(2, len(normalized_ref) * 0.5) and match_count > best_match_count:
            best_match_count = match_count
            best_row = i
            if match_count >= len(normalized_ref) * 0.8:
                return best_row
    return best_row if best_match_count >= 2 else detect_header_row_heuristic(df_raw)


def find_header_row_with_keys(df_raw: pd.DataFrame, key_columns: List[str], ws=None) -> int:
    if not key_columns:
        return detect_header_row_heuristic(df_raw, ws=ws)
    normalized_keys = {normalize_colname(k) for k in key_columns if k.strip()}
    for i in range(min(15, len(df_raw))):
        row = df_raw.iloc[i]
        row_values = {
            normalize_colname(str(val))
            for val in row
            if pd.notna(val) and str(val).strip()
        }
        if normalized_keys & row_values:
            return i
    return detect_header_row_heuristic(df_raw, ws=ws)


def detect_header_row_heuristic(df_raw: pd.DataFrame, ws=None) -> int:
    """
    Multi-signal header row detection searching the first 15 rows.
    Scores every candidate row; the highest scorer wins.

    Signals (with weights):
      bold_ratio         x3.0  — bold cells are almost always headers
      fill_ratio         x1.5  — header rows tend to be densely filled
      all_string         x1.0  — header cells are strings, data mixes types
      unique_ratio       x1.0  — header values are unique within the row
      name_like          x1.0  — values look like column names (start with letter, <40 chars)
      post_consistency   x1.0  — rows after the header should be uniform
      low_numeric        x0.5  — headers rarely contain only numbers
      row_penalty        x0.1  — prefer earlier rows when scores are equal

    Penalties:
      horizontal_merge   -2.5  — rows whose cells span multiple columns are
                                  group-label rows sitting above the real header
      vertical_cont      skip  — rows inside a vertical merge are never headers

    Skips: empty rows, single long-description rows (>50 chars, 1 cell),
           vertical-merge continuation rows.

    ws: optional openpyxl Worksheet — enables bold and merge signals.
        When None, only pandas-based signals are used.
    """
    if df_raw.empty:
        return 0

    head     = df_raw.head(15)
    n_cols   = df_raw.shape[1] or 1
    search_n = len(head)

    # ── Pre-compute merge info from openpyxl ─────────────────────────
    # horizontal_merge_rows: rows that START wide horizontal merges (group labels)
    # vertical_cont_rows:    rows that are continuations of a vertical merge
    horizontal_merge_rows: set[int] = set()
    vertical_cont_rows:    set[int] = set()
    xl_rows: list = []

    if ws is not None:
        xl_rows = list(ws.iter_rows(max_row=search_n))
        for mr in ws.merged_cells.ranges:
            ri        = mr.min_row - 1          # 0-indexed
            col_span  = mr.max_col - mr.min_col
            row_span  = mr.max_row - mr.min_row
            # Horizontal multi-column merge on a single row = group label
            if col_span >= 1 and row_span == 0 and ri < search_n:
                horizontal_merge_rows.add(ri)
            # Vertical merge — mark continuation rows (not the top-left cell row)
            if row_span >= 1:
                for r in range(mr.min_row + 1, mr.max_row + 1):
                    if r - 1 < search_n:
                        vertical_cont_rows.add(r - 1)

    best_score, best_row = -1.0, 0

    for i in range(search_n):
        # Skip rows that are inside a vertical merge (never a header)
        if i in vertical_cont_rows:
            continue

        pd_row = head.iloc[i]
        vals   = [v for v in pd_row if pd.notna(v) and str(v).strip()]
        n_vals = len(vals)

        if n_vals < 2:
            continue
        # Skip single long description rows (instructions, not headers)
        if n_vals == 1 and len(str(vals[0])) > 50:
            continue

        # ── Bold signal (requires openpyxl ws) ──────────────────────
        if xl_rows and i < len(xl_rows):
            bold_count = sum(
                1 for c in xl_rows[i]
                if c.value is not None and str(c.value).strip()
                and c.font and c.font.bold
            )
            bold_r = bold_count / n_vals
        else:
            bold_r = 0.0

        # ── Content signals ───────────────────────────────────────────
        all_str   = float(all(isinstance(v, str) for v in vals))
        unique_r  = len({str(v) for v in vals}) / n_vals
        fill_r    = n_vals / n_cols
        num_r     = sum(
            1 for v in vals
            if str(v).replace(".", "", 1).replace("-", "", 1).isdigit()
        ) / n_vals
        name_like = sum(
            1 for v in vals
            if isinstance(v, str) and len(v) < 40
            and re.match(r"^[A-Za-z]", v.strip())
        ) / n_vals

        # ── Post-row consistency (look-ahead up to 8 rows) ───────────
        lookahead = df_raw.iloc[i + 1 : i + 9]
        if len(lookahead) >= 2:
            fill_counts  = lookahead.apply(lambda r: r.notna().sum(), axis=1)
            post_consist = 1.0 / (1.0 + fill_counts.std())
        else:
            post_consist = 0.0

        # ── Composite score ───────────────────────────────────────────
        score = (
            bold_r        * 3.0
            + fill_r      * 1.5
            + all_str     * 1.0
            + unique_r    * 1.0
            + name_like   * 1.0
            + post_consist * 1.0
            + (1 - num_r) * 0.5
            - i           * 0.1     # slight penalty for later rows
        )

        # ── Merged group-label penalty ────────────────────────────────
        # Rows that start wide horizontal merges are group labels sitting
        # above the real header (e.g. "Subtotal | Sort Type | Font…").
        # Penalise heavily so the dense row beneath scores higher.
        if i in horizontal_merge_rows:
            score -= 2.5

        if score > best_score:
            best_score = score
            best_row   = i

    return best_row


# ─────────────────────────────────────────────
# FILE READING  (cached + reduced read_excel calls)
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _read_sheets_cached(
    file_bytes_raw: bytes,
    key_columns_tuple: tuple,
    header_overrides_tuple: tuple = (),  # ((sheet_name, row_idx), ...)
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, int]]:
    """
    Reads all visible sheets once, detects headers, returns DataFrames.
    Cached by raw file bytes hash – avoids re-reading on every Streamlit rerun.
    header_overrides_tuple: per-sheet manual header row (0-indexed).
    When set for a sheet, auto-detection is skipped entirely.
    """
    file_bytes = BytesIO(file_bytes_raw)
    key_columns = list(key_columns_tuple)
    header_overrides: Dict[str, int] = dict(header_overrides_tuple)

    try:
        wb = load_workbook(file_bytes, data_only=True)
        # Include all sheets: visible, hidden, and very-hidden
        # Users may need to compare hidden sheets (e.g. Lookups in OLD file)
        visible_sheets = [ws.title for ws in wb.worksheets]

        # Read ALL sheets in one pass (no per-sheet re-seek)
        file_bytes.seek(0)
        all_raw: Dict[str, pd.DataFrame] = pd.read_excel(
            file_bytes, sheet_name=None, header=None, engine="openpyxl"
        )

        sheets: Dict[str, pd.DataFrame] = {}
        header_rows: Dict[str, int] = {}

        for sh in visible_sheets:
            if sh not in all_raw:
                continue
            df_raw = all_raw[sh]

            # Detect visible columns via column_dimensions
            ws = wb[sh]
            visible_col_indices = [
                col_idx - 1
                for col_idx in range(1, ws.max_column + 1)
                if not getattr(ws.column_dimensions.get(
                    ws.cell(row=1, column=col_idx).column_letter
                ), 'hidden', False)
            ]
            if not visible_col_indices:
                visible_col_indices = list(range(df_raw.shape[1]))

            if len(visible_col_indices) < df_raw.shape[1]:
                df_raw = df_raw.iloc[:, visible_col_indices].copy()

            # Detect header row — manual override takes priority
            if sh in header_overrides:
                header_row = header_overrides[sh]
            elif key_columns:
                header_row = find_header_row_with_keys(df_raw, key_columns, ws=ws)
            else:
                header_row = detect_header_row_heuristic(df_raw, ws=ws)

            # Slice header + data in-memory (no second read_excel)
            # Strip whitespace from col names (merged cells leave trailing spaces)
            # and use consistent _N dedup to match NEW file reader
            if header_row < len(df_raw):
                raw_cols = df_raw.iloc[header_row].tolist()
                df_data  = df_raw.iloc[header_row + 1:].copy()
                seen_c: Dict[str, int] = {}
                clean_cols = []
                for i, c in enumerate(raw_cols):
                    name = str(c).strip() if pd.notna(c) else f"Unnamed_{i}"
                    if not name or name == "nan":
                        name = f"Unnamed_{i}"
                    if name in seen_c:
                        seen_c[name] += 1
                        clean_cols.append(f"{name}_{seen_c[name]}")
                    else:
                        seen_c[name] = 0
                        clean_cols.append(name)
                df_data.columns = clean_cols
                df_data = df_data.reset_index(drop=True)
            else:
                df_data = df_raw.copy()
                df_data.columns = [str(c).strip() for c in df_data.columns]

            # Flatten MultiIndex columns if present
            if isinstance(df_data.columns, pd.MultiIndex):
                df_data.columns = [
                    " ".join(str(c) for c in col if str(c) != "nan").strip()
                    for col in df_data.columns
                ]

            sheets[sh]      = df_data.astype(str)
            header_rows[sh] = header_row

        return sheets, header_rows

    except Exception:
        # Fallback: simple read_excel with header=0
        file_bytes.seek(0)
        fallback = pd.read_excel(file_bytes, sheet_name=None, engine="openpyxl")
        for sh in fallback:
            fallback[sh] = fallback[sh].astype(str)
        return fallback, {}


def read_visible_sheets_with_header_detection(
    file_bytes: BytesIO,
    key_columns: List[str] = None,
    header_overrides: Dict[str, int] = None,
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, int]]:
    raw = file_bytes.read()
    overrides_tuple = tuple(sorted((header_overrides or {}).items()))
    return _read_sheets_cached(raw, tuple(key_columns or []), overrides_tuple)


# ─────────────────────────────────────────────
# COLUMN ALIGNMENT
# ─────────────────────────────────────────────

def _normalize_for_match(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r'([_.]\d+)+$', '', s)  # strip both _1 and .1 dedup suffixes
    return re.sub(r'[^a-z0-9]', '', s)


def align_new_columns_to_reference(new_cols: List[str], ref_cols: List[str]) -> List[str]:
    new_cols = [str(c) for c in new_cols]
    ref_norm_map = {i: _normalize_for_match(c) for i, c in enumerate(ref_cols)}
    new_norm_map = {i: _normalize_for_match(c) for i, c in enumerate(new_cols)}
    assigned = set()
    renamed  = list(new_cols)

    # Exact match first
    for ref_i, ref_norm in ref_norm_map.items():
        for new_i, new_norm in new_norm_map.items():
            if new_i in assigned:
                continue
            if ref_norm and new_norm == ref_norm:
                renamed[new_i] = ref_cols[ref_i]
                assigned.add(new_i)
                break

    # Substring match fallback
    for ref_i, ref_norm in ref_norm_map.items():
        if not ref_norm:
            continue
        for new_i, new_norm in new_norm_map.items():
            if new_i in assigned:
                continue
            if ref_norm in new_norm or new_norm in ref_norm:
                renamed[new_i] = ref_cols[ref_i]
                assigned.add(new_i)
                break

    # Deduplicate
    seen: Dict[str, int] = {}
    final = []
    for name in renamed:
        if name in seen:
            seen[name] += 1
            final.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            final.append(name)
    return final


# ─────────────────────────────────────────────
# KEY SELECTION
# ─────────────────────────────────────────────

def find_best_valid_key(df1: pd.DataFrame, df2: pd.DataFrame, keys: List[str]) -> Tuple[List[str], str]:
    if not keys:
        return [], "No keys provided"

    def norm(s: str) -> str:
        s = str(s).strip().lower()
        s = re.sub(r'(_\d+)+$', '', s)    # FIX: was r'(_\\d+)+$'
        return re.sub(r'[^a-z0-9]', '', s)

    norm1 = {norm(c): c for c in df1.columns}
    norm2 = {norm(c): c for c in df2.columns}

    best_key, best_uniqueness, best_key_name = None, 0.0, None

    for k in keys:
        nk = norm(k)
        matched_col = None
        match_score = 0.0

        for n1, c1 in norm1.items():
            r1 = SequenceMatcher(None, nk, n1).ratio()
            if r1 < 0.8:
                continue
            for n2 in norm2:
                r2 = SequenceMatcher(None, nk, n2).ratio()
                if r2 < 0.8:
                    continue
                avg = (r1 + r2) / 2
                if avg > match_score:
                    match_score = avg
                    matched_col = c1

        if matched_col:
            try:
                u1 = df1[matched_col].fillna("").astype(str).nunique() / max(1, len(df1))
                u2 = df2[matched_col].fillna("").astype(str).nunique() / max(1, len(df2))
                uniqueness = (u1 + u2) / 2
                if uniqueness > best_uniqueness:
                    best_uniqueness = uniqueness
                    best_key = [matched_col]
                    best_key_name = k
            except Exception:
                continue

    if best_key:
        return best_key, f"Key: '{best_key_name}' ({best_uniqueness:.0%} unique)"
    return [], "No valid keys found"


# ─────────────────────────────────────────────
# COMPARISON FUNCTIONS
# ─────────────────────────────────────────────

def compare_key_based(df1, df2, keys):
    valid_keys, key_desc = find_best_valid_key(df1, df2, keys)
    if not valid_keys:
        raise ValueError("No valid key columns found")

    common = list(df1.columns.intersection(df2.columns))
    df1f = df1[common].fillna("").copy()
    df2f = df2[common].fillna("").copy()
    n1, n2 = normalize_df(df1f), normalize_df(df2f)

    key_str_1 = n1[valid_keys].astype(str).agg("||".join, axis=1)
    key_str_2 = n2[valid_keys].astype(str).agg("||".join, axis=1)
    n1["__key__"] = key_str_1
    n2["__key__"] = key_str_2
    df1f["__key__"] = key_str_1
    df2f["__key__"] = key_str_2

    k1, k2 = set(n1["__key__"]), set(n2["__key__"])
    common_keys  = k1 & k2
    added_keys   = k2 - k1
    removed_keys = k1 - k2

    non_key_cols = [c for c in common if c not in valid_keys and c != "__key__"]

    # Build index maps for O(1) lookup instead of repeated .loc[]
    idx1 = n1.set_index("__key__")
    idx2 = n2.set_index("__key__")
    raw1 = df1f.set_index("__key__")
    raw2 = df2f.set_index("__key__")

    changed_old, changed_new = [], []
    for key in common_keys:
        if key not in idx1.index or key not in idx2.index:
            continue
        r1, r2 = idx1.loc[key], idx2.loc[key]
        # Handle duplicate keys → take first row
        if isinstance(r1, pd.DataFrame): r1 = r1.iloc[0]
        if isinstance(r2, pd.DataFrame): r2 = r2.iloc[0]
        if any(str(r1[c]) != str(r2[c]) for c in non_key_cols):
            changed_old.append(raw1.loc[key].iloc[0] if isinstance(raw1.loc[key], pd.DataFrame) else raw1.loc[key])
            changed_new.append(raw2.loc[key].iloc[0] if isinstance(raw2.loc[key], pd.DataFrame) else raw2.loc[key])

    co = pd.DataFrame(changed_old)[common] if changed_old else pd.DataFrame(columns=common)
    cn = pd.DataFrame(changed_new)[common] if changed_new else pd.DataFrame(columns=common)
    added   = df2f.loc[df2f["__key__"].isin(added_keys),   common].reset_index(drop=True)
    removed = df1f.loc[df1f["__key__"].isin(removed_keys), common].reset_index(drop=True)

    return co.reset_index(drop=True), cn.reset_index(drop=True), added, removed, key_desc


def compare_keyless(df1, df2):
    if df1.empty and df2.empty:
        return pd.DataFrame(columns=df1.columns), pd.DataFrame(columns=df1.columns)

    n1 = normalize_df(df1)
    n2 = normalize_df(df2)

    # Guard: if no common columns exist, hashing produces meaningless results
    common = list(n1.columns.intersection(n2.columns))
    if not common:
        return pd.DataFrame(columns=df2.columns), pd.DataFrame(columns=df1.columns)

    n1c = n1[common]
    n2c = n2[common]

    # Multiset-aware comparison using value_counts instead of set isin()
    # This correctly handles duplicate rows: [A,A,B] vs [A,B,B] → 1 change
    def _row_str(df_norm: pd.DataFrame) -> pd.Series:
        return df_norm.apply(lambda r: "||".join(r.astype(str)), axis=1)

    s1 = _row_str(n1c)
    s2 = _row_str(n2c)

    counts1 = s1.value_counts()
    counts2 = s2.value_counts()

    all_keys = set(counts1.index) | set(counts2.index)
    removed_keys, added_keys = set(), set()
    for k in all_keys:
        c1 = counts1.get(k, 0)
        c2 = counts2.get(k, 0)
        if c1 > c2:
            removed_keys.add(k)
        elif c2 > c1:
            added_keys.add(k)

    added   = df2.loc[s2.isin(added_keys)].reset_index(drop=True)
    removed = df1.loc[s1.isin(removed_keys)].reset_index(drop=True)
    return added, removed


def detect_data_truncation(df1: pd.DataFrame, df2: pd.DataFrame) -> Tuple[bool, int, int]:
    old_rows = int(df1.apply(lambda r: r.notna().any(), axis=1).sum())
    new_rows = int(df2.apply(lambda r: r.notna().any(), axis=1).sum())
    threshold = max(5, old_rows * 0.1)
    return (old_rows - new_rows) >= threshold, old_rows, new_rows


# ─────────────────────────────────────────────
# PREVIEW & EXPORT
# ─────────────────────────────────────────────

def build_side_by_side_preview(changed_old: pd.DataFrame, changed_new: pd.DataFrame, keys: List[str]):
    common_cols  = changed_old.columns.intersection(changed_new.columns).tolist()
    old_norm     = normalize_df(changed_old[common_cols])
    new_norm     = normalize_df(changed_new[common_cols])

    rows = []
    for idx in range(len(changed_old)):
        row = {}
        for col in common_cols:
            row[f"{col} (Old)"] = str(changed_old.iloc[idx][col])
            row[f"{col} (New)"] = str(changed_new.iloc[idx][col])
        rows.append(row)
    combined = pd.DataFrame(rows)

    def highlight(row):
        styles = []
        ridx = row.name
        for col in combined.columns:
            if col.endswith(" (Old)"):
                base = col[:-6]
                changed = (base in old_norm.columns and
                           str(old_norm.iloc[ridx][base]) != str(new_norm.iloc[ridx][base]))
                styles.append('background-color: #3b1212; color: #fca5a5' if changed
                               else 'background-color: #0f2a1a; color: #86efac')
            elif col.endswith(" (New)"):
                base = col[:-6]
                changed = (base in new_norm.columns and
                           str(old_norm.iloc[ridx][base]) != str(new_norm.iloc[ridx][base]))
                styles.append('background-color: #0f2a1a; color: #86efac' if changed
                               else 'background-color: #0f2a1a; color: #86efac')
            else:
                styles.append('')
        return styles

    return combined.style.apply(highlight, axis=1)


def _dedup_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.columns.is_unique:
        return df
    seen: Dict[str, int] = {}
    cols = []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            cols.append(c)
    df = df.copy()
    df.columns = cols
    return df


def style_added_rows(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    df = _dedup_columns(df.copy())
    return df.style.apply(lambda r: ['background-color: #052e16; color: #86efac' for _ in r], axis=1)


def style_removed_rows(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    df = _dedup_columns(df.copy())
    return df.style.apply(lambda r: ['background-color: #450a0a; color: #fca5a5' for _ in r], axis=1)


def export_to_excel(changed_old: pd.DataFrame, changed_new: pd.DataFrame, keys: List[str]) -> BytesIO:
    wb = Workbook()
    ws_old = wb.active
    ws_old.title = "Old Values"
    ws_new = wb.create_sheet(title="New Values")
    fill_old = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_new = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    old_norm = normalize_df(changed_old)
    new_norm = normalize_df(changed_new)

    for ws, df, fill, norm_other in [
        (ws_old, changed_old, fill_old, new_norm),
        (ws_new, changed_new, fill_new, old_norm),
    ]:
        for c, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=c, value=col_name)
        for r_idx in range(len(df)):
            for c_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=r_idx + 2, column=c_idx, value=df.iloc[r_idx][col_name])
                if (col_name not in keys and col_name in old_norm.columns and
                        str(old_norm.iloc[r_idx][col_name]) != str(new_norm.iloc[r_idx][col_name])):
                    cell.fill = fill

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────
# UI LAYOUT
# ─────────────────────────────────────────────

# ── Step 1: Upload
st.markdown('<div class="step-label">Step 1 — Upload Files</div>', unsafe_allow_html=True)

# Uploader key counter — incrementing it forces Streamlit to re-mount
# the file_uploader widgets fresh, which clears all selected files.
if "upload_key" not in st.session_state:
    st.session_state.upload_key = 0

col_l, col_r, col_clr = st.columns([2, 2, 1])
with col_l:
    left = st.file_uploader(
        "OLD files", ["xlsx", "xls"], accept_multiple_files=True,
        label_visibility="visible",
        key=f"uploader_left_{st.session_state.upload_key}",
    )
with col_r:
    right = st.file_uploader(
        "NEW files", ["xlsx", "xls"], accept_multiple_files=True,
        label_visibility="visible",
        key=f"uploader_right_{st.session_state.upload_key}",
    )
with col_clr:
    st.markdown("<div style='height:1.9rem'></div>", unsafe_allow_html=True)
    if st.button("🗑 Clear files", key="clear_uploads",
                 help="Remove all uploaded files and reset the session",
                 use_container_width=True):
        st.session_state.upload_key += 1
        # Also wipe all derived state so nothing stale carries over
        for k in ["header_overrides", "_hdr_files_key", "manual_pairs",
                  "left_files", "right_files"]:
            st.session_state.pop(k, None)
        # Clear widget-level state for header inputs
        for k in list(st.session_state.keys()):
            if k.startswith("hdr_") or k.startswith("_prev_auto_"):
                del st.session_state[k]
        st.rerun()

# ── Step 2: Config
st.markdown('<div class="step-label" style="margin-top:1.5rem">Step 2 — Configure</div>',
            unsafe_allow_html=True)

cfg_col1, cfg_col2, cfg_col3 = st.columns([2, 1, 1])
with cfg_col1:
    keys_str = st.text_input("Key columns (comma-separated)", "",
                              placeholder="e.g.  ID, Employee Name, Code")
with cfg_col2:
    ignore_suffix = st.checkbox("Ignore filename suffix", True,
                                help="Strips text after the last underscore before matching files")
with cfg_col3:
    threshold = st.slider("Match threshold", 0.5, 1.0, 0.85, 0.05,
                          help="Minimum similarity score for auto-pairing files")

# ── Header Row Overrides (shown only when files uploaded)
if left or right:
    all_files = (left or []) + (right or [])

    # Run auto-detection eagerly — cached so it's instant on reruns
    auto_rows = get_auto_detected_rows(all_files)

    # Collect ordered sheet names across all files
    all_sheet_names: list[str] = []
    for sh in auto_rows:
        if sh not in all_sheet_names:
            all_sheet_names.append(sh)

    # Initialise session state on first load, or when files change
    files_key = tuple(sorted(f.name for f in all_files))
    if (st.session_state.get("_hdr_files_key") != files_key
            or "header_overrides" not in st.session_state):
        # Reset overrides when the uploaded file set changes
        st.session_state.header_overrides = {}
        st.session_state._hdr_files_key = files_key

    with st.expander("🔢 Header Row Overrides", expanded=False):
        st.markdown(
            '<div style="color:#7888a8;font-size:0.82rem;margin-bottom:0.75rem">'
            'Auto-detection is shown below. Change a value only if the detected row '
            'is wrong — for example when a sheet has a merged group-label row above '
            'the real header. <b>Row 1 = first row of the sheet.</b></div>',
            unsafe_allow_html=True,
        )

        if not all_sheet_names:
            st.caption("Upload files above to see sheet names.")
        else:
            cols_per_row = 3
            sheet_chunks = [all_sheet_names[i:i+cols_per_row]
                            for i in range(0, len(all_sheet_names), cols_per_row)]

            for chunk in sheet_chunks:
                grid = st.columns(cols_per_row)
                for col, sh in zip(grid, chunk):
                    with col:
                        auto_val = auto_rows.get(sh, 1)
                        widget_key = f"hdr_{sh}"

                        # Streamlit ignores value= after the first render — it uses
                        # the widget's own session state instead. We must seed the
                        # widget key directly so the auto-detected value is shown
                        # correctly when files change or auto-detection updates.
                        if widget_key not in st.session_state:
                            # First time this widget is rendered — seed with override
                            # (if the user already set one) or auto-detected value.
                            st.session_state[widget_key] = int(
                                st.session_state.header_overrides.get(sh, auto_val)
                            )

                        # If files changed, reset widget to new auto-detected value
                        # (files_key change already cleared header_overrides above)
                        prev_auto_key = f"_prev_auto_{sh}"
                        if st.session_state.get(prev_auto_key) != auto_val:
                            st.session_state[widget_key] = int(
                                st.session_state.header_overrides.get(sh, auto_val)
                            )
                            st.session_state[prev_auto_key] = auto_val

                        current = st.session_state[widget_key]
                        is_overridden = (current != auto_val)

                        label = (
                            f'"{sh}" ✏️'     # pencil = user-overridden
                            if is_overridden
                            else f'"{sh}" 🤖' # robot = using auto-detect
                        )
                        val = st.number_input(
                            label,
                            min_value=1, max_value=100,
                            value=int(current),
                            step=1,
                            key=widget_key,
                            help=(
                                f"Auto-detected: row {auto_val}. "
                                + ("Currently overridden." if is_overridden
                                   else "Matches auto-detection — change only if wrong.")
                            ),
                        )
                        # Only store in overrides when the user deviates from auto-detect
                        if val != auto_val:
                            st.session_state.header_overrides[sh] = val
                        elif sh in st.session_state.header_overrides:
                            # User reset back to auto-detected value → remove override
                            del st.session_state.header_overrides[sh]

            # Summary line
            n_overridden = len(st.session_state.header_overrides)
            if n_overridden:
                overridden_names = ", ".join(
                    f"{sh} → row {row}"
                    for sh, row in st.session_state.header_overrides.items()
                )
                st.markdown(
                    f'<div style="color:#fbbf24;font-size:0.78rem;margin-top:0.5rem">'
                    f'✏️ {n_overridden} sheet(s) overridden: {overridden_names}</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div style="color:#4a5878;font-size:0.78rem;margin-top:0.5rem">'
                    '🤖 All sheets using auto-detection</div>',
                    unsafe_allow_html=True,
                )

            if n_overridden and st.button("↺ Reset all to auto-detect", key="reset_overrides"):
                st.session_state.header_overrides = {}
                st.rerun()

# ── Step 3: Match & Run
if left and right:
    def comparable(n, ignore):
        n = n.rsplit(".", 1)[0]
        return n.rsplit("_", 1)[0].lower() if ignore and "_" in n else n.lower()

    # ── Auto-match by filename similarity ────
    auto_matched, usedR = [], set()
    for lf in left:
        lcmp = comparable(lf.name, ignore_suffix)
        best, ratio = None, 0.0
        for rf in right:
            if rf.name in usedR:
                continue
            r = SequenceMatcher(None, lcmp, comparable(rf.name, ignore_suffix)).ratio()
            if r >= threshold and r > ratio:
                best, ratio = rf, r
        if best:
            auto_matched.append((lf, best))
            usedR.add(best.name)

    st.markdown('<div class="step-label" style="margin-top:1.5rem">Step 3 — Review & Run</div>',
                unsafe_allow_html=True)

    # ── Manual pairing UI (always shown when uploads exist) ──────────
    left_names  = {f.name: f for f in left}
    right_names = {f.name: f for f in right}

    if not auto_matched:
        st.markdown(
            '<div class="warn-box">⚠️ No files matched automatically — the filenames are too '
            'different. Use <b>Manual Pairing</b> below to pair them directly.</div>',
            unsafe_allow_html=True,
        )

    with st.expander(
        f"🔧 Manual Pairing {'(recommended — auto-match found 0 pairs)' if not auto_matched else '(optional override)'}",
        expanded=not auto_matched,
    ):
        st.markdown(
            '<div style="color:#7888a8;font-size:0.82rem;margin-bottom:0.75rem">'
            'Select an OLD file and a NEW file to compare, then click Add Pair. '
            'Manual pairs override auto-matching for those files.</div>',
            unsafe_allow_html=True,
        )
        mp_col1, mp_col2, mp_col3 = st.columns([2, 2, 1])
        with mp_col1:
            sel_old = st.selectbox("OLD file", list(left_names.keys()), key="manual_old")
        with mp_col2:
            sel_new = st.selectbox("NEW file", list(right_names.keys()), key="manual_new")
        with mp_col3:
            st.markdown("<div style='height:1.95rem'></div>", unsafe_allow_html=True)
            add_pair = st.button("＋ Add Pair")

        if "manual_pairs" not in st.session_state:
            st.session_state.manual_pairs = []

        if add_pair:
            entry = (sel_old, sel_new)
            if entry not in st.session_state.manual_pairs:
                st.session_state.manual_pairs.append(entry)

        if st.session_state.manual_pairs:
            st.markdown("**Added manual pairs:**")
            to_remove = []
            for idx, (on, nn) in enumerate(st.session_state.manual_pairs):
                r1, r2 = st.columns([8, 1])
                with r1:
                    st.markdown(
                        f'<div class="pair-card" style="margin-bottom:0.3rem">'
                        f'<span class="fname">{on}</span>'
                        f'<span class="arrow">⟶</span>'
                        f'<span class="fname">{nn}</span>'
                        f'<span style="color:#6382ff;font-family:monospace;font-size:0.72rem">manual</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                with r2:
                    if st.button("✕", key=f"rm_{idx}"):
                        to_remove.append(idx)
            for idx in reversed(to_remove):
                st.session_state.manual_pairs.pop(idx)

        if st.button("🗑 Clear all manual pairs", key="clear_manual"):
            st.session_state.manual_pairs = []

    # ── Merge auto + manual pairs (manual takes priority) ────────────
    manual_old_names = {on for on, _ in st.session_state.get("manual_pairs", [])}
    merged_matched = [
        (lf, rf) for lf, rf in auto_matched
        if lf.name not in manual_old_names
    ]
    for old_name, new_name in st.session_state.get("manual_pairs", []):
        if old_name in left_names and new_name in right_names:
            merged_matched.append((left_names[old_name], right_names[new_name]))

    matched = merged_matched  # final list used for comparison

    # ── Pair summary ─────────────────────────────────────────────────
    if matched:
        with st.expander(f"📋 {len(matched)} pair(s) ready to compare", expanded=False):
            for lf, rf in matched:
                is_manual = lf.name in manual_old_names
                tag = '<span style="color:#6382ff;font-family:monospace;font-size:0.72rem">manual</span>' if is_manual else ''
                st.markdown(
                    f'<div class="pair-card">'
                    f'<span class="fname">{lf.name}</span>'
                    f'<span class="arrow">⟶</span>'
                    f'<span class="fname">{rf.name}</span>'
                    f'{tag}</div>',
                    unsafe_allow_html=True,
                )
        unmatched_l = [f.name for f in left  if f.name not in {lf.name for lf, _ in matched}]
        unmatched_r = [f.name for f in right if f.name not in {rf.name for _, rf in matched}]
        if unmatched_l or unmatched_r:
            st.markdown(
                f'<div class="warn-box">⚠️ Still unmatched — OLD: {", ".join(unmatched_l) or "none"} '
                f'· NEW: {", ".join(unmatched_r) or "none"}</div>',
                unsafe_allow_html=True,
            )

    run_clicked = st.button(
        "▶  Run Comparison",
        disabled=not matched,
        use_container_width=False,
    )

    if run_clicked and matched:
        prog = st.progress(0, text="Starting…")
        keys = [k.strip() for k in keys_str.split(",") if k.strip()]
        files_with_changes = []

        for i, (lf, rf) in enumerate(matched):
            prog.progress((i) / len(matched), text=f"Comparing {lf.name} …")
            try:
                decL = decrypt_file(lf)
                decR = decrypt_file(rf)

                # Apply manual header overrides (convert 1-based UI input to 0-based index)
                header_overrides = {
                    sh: (row - 1)
                    for sh, row in st.session_state.get("header_overrides", {}).items()
                    if row is not None
                }
                shL, header_rows_old = read_visible_sheets_with_header_detection(
                    decL, key_columns=keys, header_overrides=header_overrides
                )

                # Read NEW file through the SAME cached pipeline as OLD
                # so it gets its own independent auto-detection with style signals.
                # Override priority: user override > NEW's own detection > OLD's detection
                shR, header_rows_new = read_visible_sheets_with_header_detection(
                    decR, key_columns=keys, header_overrides=header_overrides
                )

                # Align NEW column names to OLD column names for every shared sheet
                for sh in list(shR.keys()):
                    if sh in shL:
                        try:
                            shR[sh].columns = align_new_columns_to_reference(
                                list(shR[sh].columns), list(shL[sh].columns)
                            )
                        except Exception:
                            pass

                # Determine which sheets are missing in each file
                missing_in_new: List[str] = [sh for sh in shL if sh not in shR]
                missing_in_old: List[str] = [sh for sh in shR if sh not in shL]

            except Exception as e:
                import traceback
                st.error(f"❌ Pair {i+1}: Failed to read — {e}")
                st.code(traceback.format_exc())
                prog.progress((i + 1) / len(matched))
                continue

            changed_sheets = []
            for sh in sorted(set(shL) & set(shR)):
                d1, d2 = shL[sh], shR[sh]
                use_key = False
                key_desc = ""
                if keys:
                    try:
                        valid_keys, key_desc = find_best_valid_key(d1, d2, keys)
                        use_key = bool(valid_keys)
                    except Exception:
                        use_key = False

                try:
                    if use_key:
                        co, cn, add, rem, key_desc = compare_key_based(d1, d2, keys)
                    else:
                        co = cn = pd.DataFrame()
                        add, rem = compare_keyless(d1, d2)
                except Exception:
                    co = cn = pd.DataFrame()
                    add, rem = compare_keyless(d1, d2)

                is_trunc, old_rows, new_rows = detect_data_truncation(d1, d2)

                if not (cn.empty and add.empty and rem.empty):
                    changed_sheets.append((sh, co, cn, add, rem, use_key, key_desc,
                                           is_trunc, old_rows, new_rows))

            # Always record the pair so missing-sheet notices are shown
            files_with_changes.append((i + 1, lf.name, rf.name, changed_sheets,
                                       missing_in_new, missing_in_old))

            prog.progress((i + 1) / len(matched), text=f"Done {i+1}/{len(matched)}")

        prog.empty()

        # ── Results ──────────────────────────────
        truly_identical = all(
            not cs and not mn and not mo
            for _, _, _, cs, mn, mo in files_with_changes
        ) if files_with_changes else False

        if len(matched) == 0:
            st.markdown(
                '<div class="warn-box">⚠️ No pairs were compared — '
                'use Manual Pairing above to pair files with different names.</div>',
                unsafe_allow_html=True,
            )
        elif truly_identical:
            st.markdown(
                f'<div class="success-box">✅ All {len(matched)} file pair(s) are identical.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div class="metric-row">'
                f'<div class="metric-pill info"><span class="val">{len(matched)}</span> pairs compared</div>'
                f'<div class="metric-pill changed"><span class="val">{len(files_with_changes)}</span> with differences</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            for file_num, old_name, new_name, changed_sheets, missing_in_new, missing_in_old in files_with_changes:
                st.markdown(f"### 📂 Pair {file_num} of {len(matched)}")
                st.markdown(
                    f'<div class="pair-card" style="margin-bottom:1rem">'
                    f'<span class="fname">{old_name}</span>'
                    f'<span class="arrow">⟶</span>'
                    f'<span class="fname">{new_name}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # ── Missing sheet notices ─────────────────────
                if missing_in_new:
                    sheets_list = ", ".join(f"<b>{s}</b>" for s in missing_in_new)
                    st.markdown(
                        f'<div class="warn-box">📋 Sheet(s) only in OLD file (skipped): {sheets_list}</div>',
                        unsafe_allow_html=True,
                    )
                if missing_in_old:
                    sheets_list = ", ".join(f"<b>{s}</b>" for s in missing_in_old)
                    st.markdown(
                        f'<div class="info-box">📋 Sheet(s) only in NEW file (skipped): {sheets_list}</div>',
                        unsafe_allow_html=True,
                    )
                if not changed_sheets and not missing_in_new and not missing_in_old:
                    st.markdown(
                        '<div class="success-box" style="text-align:left;margin-bottom:1rem">'
                        '✅ All common sheets are identical.</div>',
                        unsafe_allow_html=True,
                    )

                for sh, co, cn, add, rem, use_key, key_desc, is_trunc, old_rows, new_rows in changed_sheets:
                    parts = []
                    if len(cn)  > 0: parts.append(f"{len(cn)} changed")
                    if len(add) > 0: parts.append(f"{len(add)} added")
                    if len(rem) > 0: parts.append(f"{len(rem)} removed")
                    summary = " · ".join(parts) or "differences detected"

                    st.markdown(
                        f'<div class="sheet-header">'
                        f'<span class="sheet-name">📄 {sh}</span>'
                        f'<span class="sheet-summary">{summary}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                    # Metric pills
                    pills = ""
                    if len(cn)  > 0: pills += f'<div class="metric-pill changed"><span class="val">{len(cn)}</span> changed</div>'
                    if len(add) > 0: pills += f'<div class="metric-pill added"><span class="val">{len(add)}</span> added</div>'
                    if len(rem) > 0: pills += f'<div class="metric-pill removed"><span class="val">{len(rem)}</span> removed</div>'
                    if pills:
                        st.markdown(f'<div class="metric-row">{pills}</div>', unsafe_allow_html=True)

                    if is_trunc:
                        st.markdown(
                            f'<div class="warn-box">⚠️ <b>Truncation detected</b> — '
                            f'OLD: {old_rows} rows · NEW: {new_rows} rows '
                            f'({old_rows - new_rows} rows missing)</div>',
                            unsafe_allow_html=True,
                        )

                    if key_desc and "Key:" in key_desc:
                        st.markdown(
                            f'<span class="diff-label key">🔑 {key_desc}</span>',
                            unsafe_allow_html=True,
                        )

                    if not cn.empty and use_key:
                        st.markdown('<span class="diff-label changed">🔄 Changed rows</span>',
                                    unsafe_allow_html=True)
                        st.dataframe(build_side_by_side_preview(co, cn, keys), use_container_width=True)
                        st.download_button(
                            "📥 Download changes (.xlsx)",
                            export_to_excel(co, cn, keys),
                            file_name=f"Pair{file_num}_{sh}_changes.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{file_num}_{sh}",
                        )
                    elif not cn.empty:
                        st.markdown('<span class="diff-label changed">🔄 Changed rows</span>',
                                    unsafe_allow_html=True)
                        st.dataframe(cn, use_container_width=True)

                    if not add.empty:
                        st.markdown('<span class="diff-label added">➕ Added rows</span>',
                                    unsafe_allow_html=True)
                        st.dataframe(style_added_rows(add), use_container_width=True)

                    if not rem.empty:
                        st.markdown('<span class="diff-label removed">➖ Removed rows</span>',
                                    unsafe_allow_html=True)
                        st.dataframe(style_removed_rows(rem), use_container_width=True)

                    st.divider()

else:
    st.markdown(
        '<div style="color:#475569;font-size:0.9rem;padding:1.5rem 0;">Upload OLD and NEW files above to begin.</div>',
        unsafe_allow_html=True,
    )
